"""
Bitrix24 - CRM Report App
Halaman 1: Deals Belum Invoice
Halaman 2: Outstanding Qty
Role: super_admin, PKR, PKL
"""
st.session_state.clear()
import streamlit as st
import requests
import pandas as pd
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO
from collections import defaultdict
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from datetime import date, timedelta
from auth import login_page_otp

# ==================== CONFIG ====================
WEBHOOK    = st.secrets["config"]["BITRIX_WEBHOOK"]
DELAY      = 0.1

SMTP_HOST  = st.secrets["config"]["SMTP_HOST"]
SMTP_PORT  = int(st.secrets["config"]["SMTP_PORT"])
SMTP_USER  = st.secrets["config"]["SMTP_USER"]
SMTP_PASS  = st.secrets["config"]["SMTP_PASS"]
EMAIL_FROM = st.secrets["config"]["EMAIL_FROM"]
CC_EMAILS  = [st.secrets["config"]["CC_EMAIL"]]
# ================================================




def make_session():
    session = requests.Session()
    retry = Retry(total=5, backoff_factor=2, status_forcelist=[500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retry))
    return session

SESSION = make_session()


def bx_get(method, params, timeout=60):
    url = WEBHOOK + method
    for attempt in range(1, 4):
        try:
            resp = SESSION.get(url, params=params, timeout=timeout)
            resp.raise_for_status()
            time.sleep(DELAY)
            return resp.json()
        except requests.exceptions.ConnectTimeout:
            if attempt == 3:
                st.session_state["fetch_error"] = "timeout"
            time.sleep(attempt * 10)
        except requests.exceptions.RequestException as e:
            st.session_state["fetch_error"] = str(e)
            break
    return {}


def bx_post(method, params, timeout=60):
    url = WEBHOOK + method
    for attempt in range(1, 4):
        try:
            resp = SESSION.post(url, json=params, timeout=timeout)
            resp.raise_for_status()
            time.sleep(DELAY)
            return resp.json()
        except requests.exceptions.ConnectTimeout:
            if attempt == 3:
                st.session_state["fetch_error"] = "timeout"
            time.sleep(attempt * 10)
        except requests.exceptions.RequestException as e:
            st.session_state["fetch_error"] = str(e)
            break
    return {}


def bitrix_get_all(method, params):
    results = []
    start = 0
    while True:
        p = {**params, "start": start}
        data = bx_get(method, p)
        batch = data.get("result", [])
        results.extend(batch)
        if "next" not in data:
            break
        start = data["next"]
    return results


def get_users_batch(user_ids):
    unique_ids = list(set(str(uid) for uid in user_ids if uid))
    user_map = {}
    chunk_size = 50
    for i in range(0, len(unique_ids), chunk_size):
        chunk = unique_ids[i:i+chunk_size]
        params = {"select[]": ["ID", "NAME", "LAST_NAME"]}
        for j, uid in enumerate(chunk):
            params[f"filter[ID][{j}]"] = uid
        data = bx_get("user.get.json", params)
        for u in data.get("result", []):
            full_name = f"{u.get('NAME', '')} {u.get('LAST_NAME', '')}".strip()
            user_map[str(u["ID"])] = full_name or str(u["ID"])
    return user_map


def get_companies_batch(company_ids):
    unique_ids = list(set(str(cid) for cid in company_ids if cid))
    company_map = {}
    chunk_size = 50
    for i in range(0, len(unique_ids), chunk_size):
        chunk = unique_ids[i:i+chunk_size]
        params = {"select[]": ["ID", "TITLE", "COMPANY_TYPE"]}
        for j, cid in enumerate(chunk):
            params[f"filter[ID][{j}]"] = cid
        data = bx_get("crm.company.list.json", params)
        for c in data.get("result", []):
            company_map[str(c["ID"])] = {
                "name": c.get("TITLE", "-"),
                "type": c.get("COMPANY_TYPE", "")
            }
    return company_map


# ==================== ROLE HELPER ====================


def get_allowed_company_types(role):
    role_types = st.secrets.get("role_company_types", {})
    allowed = role_types.get(role, [])
    if role == "super_admin":
        return None
    return list(allowed)


def filter_by_company_type(deals_data, company_map, allowed_types):
    if allowed_types is None:
        return deals_data
    filtered = []
    for d in deals_data:
        company_id = str(d.get("COMPANY_ID", ""))
        company_info = company_map.get(company_id, {})
        company_type = company_info.get("type", "")
        if company_type in allowed_types:
            filtered.append(d)
    return filtered


# ==================== PAGE 1: DEALS BELUM INVOICE ====================
def fetch_deals_belum_invoice(start_date, end_date, allowed_types):
    progress = st.progress(0)
    status = st.empty()

    status.text("📋 Mengambil deals WON...")
    all_deals = bitrix_get_all(
        "crm.deal.list.json",
        {
            "filter[STAGE_SEMANTIC_ID]": "S",
            "filter[>=CLOSEDATE]": start_date.strftime("%Y-%m-%d"),
            "filter[<=CLOSEDATE]": end_date.strftime("%Y-%m-%d"),
            "select[]": ["ID", "TITLE", "STAGE_ID", "OPPORTUNITY",
                         "ASSIGNED_BY_ID", "COMPANY_ID", "CLOSEDATE"]
        }
    )
    progress.progress(20)
    status.text(f"✅ {len(all_deals)} deals WON ditemukan")

    status.text("🏢 Mengambil info company...")
    all_company_ids = [d.get("COMPANY_ID") for d in all_deals]
    company_map = get_companies_batch(all_company_ids)
    progress.progress(35)

    all_deals = filter_by_company_type(all_deals, company_map, allowed_types)

    status.text("📄 Mengambil data invoice...")
    all_invoices_raw = bitrix_get_all(
        "crm.invoice.list.json",
        {"select": ["UF_DEAL_ID"]}
    )
    invoice_deal_ids = set()
    for inv in all_invoices_raw:
        deal_id = str(inv.get("UF_DEAL_ID", ""))
        if deal_id and deal_id != "0":
            invoice_deal_ids.add(deal_id)
    progress.progress(60)

    deals_filtered = [d for d in all_deals if str(d["ID"]) not in invoice_deal_ids]

    status.text("👤 Mengambil info user...")
    all_user_ids = [d.get("ASSIGNED_BY_ID") for d in deals_filtered]
    user_map = get_users_batch(all_user_ids)
    progress.progress(80)

    rows = []
    for deal in deals_filtered:
        user_id      = str(deal.get("ASSIGNED_BY_ID", ""))
        company_id   = str(deal.get("COMPANY_ID", ""))
        closedate    = deal.get("CLOSEDATE", "")[:10] if deal.get("CLOSEDATE") else ""
        company_info = company_map.get(company_id, {"name": "-", "type": "-"})
        rows.append({
            "Deal ID":      deal.get("ID"),
            "Deal Name":    deal.get("TITLE", ""),
            "Stage":        deal.get("STAGE_ID", ""),
            "Amount":       float(deal.get("OPPORTUNITY", 0) or 0),
            "Responsible":  user_map.get(user_id, user_id),
            "End Date":     closedate,
            "Company ID":   company_id or "-",
            "Company Name": company_info.get("name", "-"),
            "Company Type": company_info.get("type", "-"),
        })

    progress.progress(100)
    status.text(f"✅ Selesai! {len(rows)} deals belum invoice.")
    time.sleep(0.5)
    status.empty()
    progress.empty()

    return pd.DataFrame(rows)


# ==================== PAGE 2: OUTSTANDING QTY ====================
def fetch_outstanding_qty(start_date, end_date, allowed_types):
    progress = st.progress(0)
    status = st.empty()

    status.text("📋 Mengambil deals WON...")
    all_deals = bitrix_get_all(
        "crm.deal.list.json",
        {
            "filter[STAGE_SEMANTIC_ID]": "S",
            "filter[>=CLOSEDATE]": start_date.strftime("%Y-%m-%d"),
            "filter[<=CLOSEDATE]": end_date.strftime("%Y-%m-%d"),
            "select[]": ["ID", "TITLE", "STAGE_ID", "OPPORTUNITY",
                         "ASSIGNED_BY_ID", "COMPANY_ID", "CLOSEDATE"]
        }
    )
    progress.progress(15)
    status.text(f"✅ {len(all_deals)} deals WON ditemukan")

    if not all_deals:
        progress.empty()
        status.empty()
        return pd.DataFrame()

    status.text("🏢 Mengambil info company...")
    all_company_ids = [d.get("COMPANY_ID") for d in all_deals]
    company_map = get_companies_batch(all_company_ids)
    progress.progress(25)

    all_deals = filter_by_company_type(all_deals, company_map, allowed_types)
    if not all_deals:
        progress.empty()
        status.empty()
        return pd.DataFrame()

    deal_ids_set = set(str(d["ID"]) for d in all_deals)

    status.text("📄 Mengambil data invoice...")
    all_invoices_raw = bitrix_get_all(
        "crm.invoice.list.json",
        {"select": ["ID", "UF_DEAL_ID", "ACCOUNT_NUMBER", "DATE_BILL", "STATUS_ID", "PRICE"]}
    )
    progress.progress(40)

    invoice_map = defaultdict(list)
    invoice_amount_map = defaultdict(float)
    for inv in all_invoices_raw:
        deal_id = str(inv.get("UF_DEAL_ID", ""))
        if deal_id in deal_ids_set and inv.get("STATUS_ID") != "D":
            invoice_map[deal_id].append({
                "id":     inv.get("ID"),
                "number": inv.get("ACCOUNT_NUMBER", inv.get("ID", "")),
                "date":   inv.get("DATE_BILL", "")[:10] if inv.get("DATE_BILL") else ""
            })
            invoice_amount_map[deal_id] += float(inv.get("PRICE", 0) or 0)

    deals_to_process = []
    for d in all_deals:
        deal_id     = str(d["ID"])
        deal_amount = float(d.get("OPPORTUNITY", 0) or 0)
        inv_amount  = invoice_amount_map.get(deal_id, 0)
        if inv_amount < deal_amount:
            deals_to_process.append(d)

    status.text(f"📊 {len(deals_to_process)} deals perlu diproses")
    progress.progress(45)

    if not deals_to_process:
        progress.empty()
        status.empty()
        return pd.DataFrame()

    status.text("📦 Mengambil product rows invoice...")
    inv_product_map = defaultdict(list)
    all_invoice_ids = list(set(
        str(inv["id"])
        for invs in invoice_map.values()
        for inv in invs
    ))
    for i, inv_id in enumerate(all_invoice_ids):
        data = bx_post("crm.productrow.list", {
            "filter": {"OWNER_TYPE": "I", "OWNER_ID": int(inv_id)},
            "select": ["OWNER_ID", "PRODUCT_ID", "PRODUCT_NAME", "QUANTITY"]
        })
        for p in (data.get("result", []) or []):
            inv_product_map[str(inv_id)].append(p)
        if (i + 1) % 50 == 0 or (i + 1) == len(all_invoice_ids):
            pct = 45 + int((i + 1) / len(all_invoice_ids) * 20)
            progress.progress(min(pct, 65))
            status.text(f"📦 Mengambil product rows invoice... ({i+1}/{len(all_invoice_ids)})")

    progress.progress(65)

    status.text("👤 Mengambil info user...")
    all_user_ids = [d.get("ASSIGNED_BY_ID") for d in deals_to_process]
    user_map = get_users_batch(all_user_ids)
    progress.progress(70)

    status.text(f"📦 Mengambil product rows deals... (0/{len(deals_to_process)})")
    deal_product_map = defaultdict(list)
    for i, deal in enumerate(deals_to_process):
        deal_id = str(deal["ID"])
        data = bx_post("crm.deal.productrows.get", {"id": deal["ID"]})
        products = data.get("result", [])
        if isinstance(products, list):
            deal_product_map[deal_id] = products
        if (i + 1) % 10 == 0 or (i + 1) == len(deals_to_process):
            status.text(f"📦 Mengambil product rows deals... ({i+1}/{len(deals_to_process)})")
            pct = 70 + int((i + 1) / len(deals_to_process) * 25)
            progress.progress(min(pct, 95))

    progress.progress(95)

    status.text("🔨 Membangun data...")
    rows = []
    for deal in deals_to_process:
        deal_id     = str(deal["ID"])
        user_id     = str(deal.get("ASSIGNED_BY_ID", ""))
        closedate   = deal.get("CLOSEDATE", "")[:10] if deal.get("CLOSEDATE") else ""
        responsible = user_map.get(user_id, user_id)

        inv_list = invoice_map.get(deal_id, [])
        invoice_label = ", ".join(
            f"{inv['number']} ({inv['date']})" if inv['date'] else inv['number']
            for inv in inv_list
        ) if inv_list else "-"

        invoiced_qty = defaultdict(float)
        for inv in inv_list:
            for p in inv_product_map.get(str(inv["id"]), []):
                key = p.get("PRODUCT_NAME", "").strip()
                invoiced_qty[key] += float(p.get("QUANTITY", 0))

        deal_products = deal_product_map.get(deal_id, [])
        if not deal_products:
            continue

        for p in deal_products:
            key               = p.get("PRODUCT_NAME", "").strip()
            qty_ordered       = float(p.get("QUANTITY", 0))
            qty_invoiced      = invoiced_qty.get(key, 0)
            outstanding       = qty_ordered - qty_invoiced

            if outstanding <= 0:
                continue

            unit_price        = float(p.get("PRICE", 0))
            outstanding_value = outstanding * unit_price

            rows.append({
                "Deal ID":           deal["ID"],
                "Deal Name":         deal.get("TITLE", ""),
                "Stage":             deal.get("STAGE_ID", ""),
                "Amount":            float(deal.get("OPPORTUNITY", 0) or 0),
                "Responsible":       responsible,
                "End Date":          closedate,
                "Product Name":      p.get("PRODUCT_NAME", ""),
                "Unit":              p.get("MEASURE_NAME", ""),
                "Unit Price":        unit_price,
                "Qty Ordered":       qty_ordered,
                "Qty Invoiced":      qty_invoiced,
                "Outstanding Qty":   outstanding,
                "Outstanding Value": outstanding_value,
                "Invoices":          invoice_label,
            })

    progress.progress(100)
    status.text(f"✅ Selesai! {len(rows)} baris data ditemukan.")
    time.sleep(0.5)
    status.empty()
    progress.empty()

    return pd.DataFrame(rows)


# ==================== EXCEL ====================
def build_excel_belum_invoice(df):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Deals Belum Invoice")
        _style_sheet(writer.sheets["Deals Belum Invoice"], currency_cols=["Amount"])
    buffer.seek(0)
    return buffer.read()


def build_excel_outstanding(df):
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Outstanding Qty")
        _style_sheet(
            writer.sheets["Outstanding Qty"],
            currency_cols=["Unit Price", "Outstanding Value", "Amount"],
            number_cols=["Qty Ordered", "Qty Invoiced", "Outstanding Qty"]
        )
        if not df.empty:
            summary = {
                "Total Deals":             [df["Deal ID"].nunique()],
                "Total Outstanding Qty":   [df["Outstanding Qty"].sum()],
                "Total Outstanding Value": [df["Outstanding Value"].sum()],
            }
            pd.DataFrame(summary).to_excel(writer, index=False, sheet_name="Summary")
    buffer.seek(0)
    return buffer.read()


def _style_sheet(ws, currency_cols=[], number_cols=[]):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Arial", size=10)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill_color = "FFFFFF" if row_idx % 2 == 0 else "EBF3FB"
        fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
        for cell in row:
            cell.font = data_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    col_map = {cell.value: cell.column for cell in ws[1]}
    for col_name in currency_cols:
        if col_name in col_map:
            col_letter = get_column_letter(col_map[col_name])
            for cell in ws[f"{col_letter}2":f"{col_letter}{ws.max_row}"]:
                for c in cell:
                    c.number_format = "#,##0.00"
    for col_name in number_cols:
        if col_name in col_map:
            col_letter = get_column_letter(col_map[col_name])
            for cell in ws[f"{col_letter}2":f"{col_letter}{ws.max_row}"]:
                for c in cell:
                    c.number_format = "#,##0"

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"


# ==================== EMAIL ====================
def send_email(to_email, excel_data, subject, summary_html):
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = EMAIL_FROM
    msg["To"]      = to_email
    if CC_EMAILS:
        msg["Cc"] = ", ".join(CC_EMAILS)

    body = MIMEMultipart("alternative")
    html = f"""
    <html><body>
    <p>Halo,</p>
    {summary_html}
    <p>Detail lengkap dapat dilihat pada file terlampir.</p>
    <br><p>Terima kasih.</p>
    </body></html>
    """
    body.attach(MIMEText(html, "html"))
    msg.attach(body)

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_data)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{subject}.xlsx"')
    msg.attach(part)

    recipients = [to_email] + CC_EMAILS
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(EMAIL_FROM, recipients, msg.as_string())


# ==================== STREAMLIT UI ====================
st.set_page_config(page_title="Bitrix24 CRM Report", page_icon="📊", layout="wide")

user_email    = st.session_state.get("user_email", "")
user_role     = st.session_state.get("user_role", "PKR")
allowed_types = get_allowed_company_types(user_role)

with st.sidebar:
    st.title("📊 Bitrix24 Report")
    st.write(f"👤 {user_email}")
    st.caption(f"Role: **{user_role}**")
    st.divider()
    page = st.radio("Pilih Halaman", [
        "📋 Deals Belum Invoice",
        "📦 Outstanding Qty"
    ])
    st.divider()
    if st.button("Logout"):
        st.session_state.clear()
        st.rerun()

# ==================== PAGE 1 ====================
if page == "📋 Deals Belum Invoice":
    st.title("📋 Deals Belum Invoice")
    st.caption("Deals WON yang belum dibuatkan invoice sama sekali")

    col_s, col_e = st.columns(2)
    with col_s:
        start_date = st.date_input(
            "Start Date (End Date)",
            value=st.session_state.get("bi_start", date.today() - timedelta(days=30)),
            key="bi_start"
        )
    with col_e:
        end_date = st.date_input(
            "End Date (End Date)",
            value=st.session_state.get("bi_end", date.today()),
            key="bi_end"
        )

    if st.button("🔄 Ambil Data", type="primary"):
        if start_date > end_date:
            st.error("Start Date tidak boleh lebih besar dari End Date!")
        else:
            st.session_state.pop("fetch_error", None)
            st.session_state.pop("df_belum_invoice", None)
            df = fetch_deals_belum_invoice(start_date, end_date, allowed_types)
            if st.session_state.get("fetch_error") == "timeout":
                st.error("""
⏱️ **Koneksi ke Bitrix24 timeout**

Proses pengambilan data terputus karena koneksi terlalu lama.
Silakan coba lagi beberapa saat kemudian.

Tips:
- Coba perkecil rentang tanggal
- Pastikan koneksi internet stabil
- Jika masalah berlanjut, hubungi IT Support
""")
            else:
                st.session_state["df_belum_invoice"]    = df
                st.session_state["excel_belum_invoice"] = build_excel_belum_invoice(df)
                st.session_state["bi_filename"] = f"Deals_Belum_Invoice_{start_date}_{end_date}.xlsx"
                st.rerun()

    if "df_belum_invoice" in st.session_state:
        df         = st.session_state["df_belum_invoice"]
        excel_data = st.session_state["excel_belum_invoice"]
        filename   = st.session_state.get("bi_filename", "Deals_Belum_Invoice.xlsx")

        if df.empty:
            st.warning(f"""
⚠️ **Tidak ada data deals yang belum dibuatkan invoice**

Tidak ditemukan deals WON yang belum memiliki invoice pada periode:
**{st.session_state.get('bi_start', '-')}** s/d **{st.session_state.get('bi_end', '-')}**

Kemungkinan penyebab:
- Semua deals pada periode ini sudah dibuatkan invoice
- Tidak ada deals yang WON pada periode tersebut
- Coba perluas rentang tanggal pencarian
""")
        else:
            col1, col2 = st.columns(2)
            col1.metric("Total Deals", len(df))
            col2.metric("Total Amount", f"{df['Amount'].sum():,.2f}")

            st.divider()
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.divider()

            col_dl, col_email = st.columns([1, 2])
            with col_dl:
                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_email:
                with st.form("form_email_belum_invoice"):
                    to_email  = st.text_input("📧 Kirim ke (TO)", placeholder="email@domain.com")
                    submitted = st.form_submit_button("📤 Kirim Email", type="primary")
                    if submitted:
                        if not to_email:
                            st.error("Email tujuan tidak boleh kosong!")
                        else:
                            try:
                                with st.spinner("Mengirim email..."):
                                    send_email(
                                        to_email, excel_data,
                                        f"Deals Belum Invoice {start_date} sd {end_date}",
                                        f"<p>Terdapat <b>{len(df)} deals</b> WON yang belum dibuatkan invoice pada periode <b>{start_date}</b> s/d <b>{end_date}</b>.</p><p>Total Amount: <b>{df['Amount'].sum():,.2f}</b></p>"
                                    )
                                st.success(f"✅ Email berhasil dikirim ke {to_email}")
                            except Exception as e:
                                st.error(f"❌ Gagal kirim email: {e}")

# ==================== PAGE 2 ====================
elif page == "📦 Outstanding Qty":
    st.title("📦 Outstanding Qty")
    st.caption("Deals WON dengan qty yang belum sepenuhnya diinvoice")

    col_s, col_e = st.columns(2)
    with col_s:
        start_date = st.date_input(
            "Start Date (End Date)",
            value=st.session_state.get("os_start", date.today() - timedelta(days=30)),
            key="os_start"
        )
    with col_e:
        end_date = st.date_input(
            "End Date (End Date)",
            value=st.session_state.get("os_end", date.today()),
            key="os_end"
        )

    if st.button("🔄 Ambil Data", type="primary"):
        if start_date > end_date:
            st.error("Start Date tidak boleh lebih besar dari End Date!")
        else:
            st.session_state.pop("fetch_error", None)
            st.session_state.pop("df_outstanding", None)
            df = fetch_outstanding_qty(start_date, end_date, allowed_types)
            if st.session_state.get("fetch_error") == "timeout":
                st.error("""
⏱️ **Koneksi ke Bitrix24 timeout**

Proses pengambilan data terputus karena koneksi terlalu lama.
Silakan coba lagi beberapa saat kemudian.

Tips:
- Coba perkecil rentang tanggal
- Pastikan koneksi internet stabil
- Jika masalah berlanjut, hubungi IT Support
""")
            else:
                st.session_state["df_outstanding"]    = df
                st.session_state["excel_outstanding"] = build_excel_outstanding(df)
                st.session_state["os_filename"] = f"Invoice_Outstanding_{start_date}_{end_date}.xlsx"
                st.rerun()

    if "df_outstanding" in st.session_state:
        df         = st.session_state["df_outstanding"]
        excel_data = st.session_state["excel_outstanding"]
        filename   = st.session_state.get("os_filename", "Invoice_Outstanding.xlsx")

        if df.empty:
            st.warning(f"""
⚠️ **Tidak ada invoice yang masih outstanding**

Tidak ditemukan outstanding qty pada periode:
**{st.session_state.get('os_start', '-')}** s/d **{st.session_state.get('os_end', '-')}**

Kemungkinan penyebab:
- Semua qty pada deals periode ini sudah sepenuhnya diinvoice
- Tidak ada deals yang WON pada periode tersebut
- Coba perluas rentang tanggal pencarian
""")
        else:
            col1, col2, col3 = st.columns(3)
            col1.metric("Total Deals", df["Deal ID"].nunique())
            col2.metric("Total Outstanding Qty", f"{df['Outstanding Qty'].sum():,.0f}")
            col3.metric("Total Outstanding Value", f"{df['Outstanding Value'].sum():,.2f}")

            st.divider()
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.divider()

            col_dl, col_email = st.columns([1, 2])
            with col_dl:
                st.download_button(
                    label="⬇️ Download Excel",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_email:
                with st.form("form_email_outstanding"):
                    to_email  = st.text_input("📧 Kirim ke (TO)", placeholder="email@domain.com")
                    submitted = st.form_submit_button("📤 Kirim Email", type="primary")
                    if submitted:
                        if not to_email:
                            st.error("Email tujuan tidak boleh kosong!")
                        else:
                            try:
                                with st.spinner("Mengirim email..."):
                                    send_email(
                                        to_email, excel_data,
                                        f"Invoice Outstanding {start_date} sd {end_date}",
                                        f"<p>Terdapat <b>{df['Deal ID'].nunique()} deals</b> dengan outstanding qty pada periode <b>{start_date}</b> s/d <b>{end_date}</b>.</p><p>Total Outstanding Value: <b>{df['Outstanding Value'].sum():,.2f}</b></p>"
                                    )
                                st.success(f"✅ Email berhasil dikirim ke {to_email}")
                            except Exception as e:
                                st.error(f"❌ Gagal kirim email: {e}")
